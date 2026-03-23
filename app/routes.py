import io
import json
import os
import threading
from datetime import datetime, timedelta

from flask import Blueprint, jsonify, request, current_app, send_from_directory, send_file
from werkzeug.utils import secure_filename

from .models import db, Product, Box, Order, OrderItem, StockLog, ManualSession, ManualTriggerLog, Settings
from . import servo_controller as servo

bp = Blueprint("main", __name__)

ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "gif", "webp"}


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


# ─── Сторінки ────────────────────────────────────────────────────────────────

@bp.route("/")
def operator_page():
    from flask import render_template
    return render_template("operator.html")


@bp.route("/admin")
def admin_page():
    from flask import render_template
    return render_template("admin.html")


@bp.route("/report")
def report_page():
    from flask import render_template
    return render_template("report.html")


# ─── API: Статус ──────────────────────────────────────────────────────────────

@bp.route("/api/status")
def get_status():
    return jsonify({
        "simulation": servo.is_simulation(),
        "i2c_available": not servo.is_simulation()
    })


# ─── API: Замовлення ──────────────────────────────────────────────────────────

@bp.route("/api/orders")
def get_orders():
    """Активні замовлення: одне picking або до 5 pending."""
    picking = Order.query.filter_by(status="picking").first()
    if picking:
        return jsonify([picking.to_dict()])
    orders = Order.query.filter_by(status="pending") \
        .order_by(Order.created_at).limit(5).all()
    return jsonify([o.to_dict() for o in orders])


@bp.route("/api/orders/all_pending")
def get_all_pending_orders():
    """Всі pending замовлення для вкладки «Замовлення до Комплектації»."""
    orders = Order.query.filter_by(status="pending").order_by(Order.created_at).all()
    return jsonify([o.to_dict() for o in orders])


@bp.route("/api/orders/completed")
def get_completed_orders():
    """Виконані замовлення (packed/done), останні 50."""
    orders = Order.query.filter(
        Order.status.in_(["packed", "done"])
    ).order_by(Order.packed_at.desc(), Order.started_at.desc()).limit(50).all()
    return jsonify([o.to_dict() for o in orders])


@bp.route("/api/orders/load", methods=["POST"])
def load_orders_from_json():
    """Завантажити замовлення з JSON файлу."""
    json_path = os.path.join(
        os.path.dirname(os.path.dirname(__file__)), "data", "orders_test.json"
    )
    if not os.path.exists(json_path):
        return jsonify({"error": "JSON файл не знайдено"}), 404

    with open(json_path, encoding="utf-8") as f:
        orders_data = json.load(f)

    added = 0
    for od in orders_data:
        if Order.query.filter_by(order_id=od["order_id"]).first():
            continue
        order = Order(
            order_id=od["order_id"],
            order_date=od.get("order_date", ""),
        )
        db.session.add(order)
        db.session.flush()
        for item in od["items"]:
            oi = OrderItem(
                order_db_id=order.id,
                articl=item["articl"],
                quantity_ordered=item["quantity"],
            )
            db.session.add(oi)
        added += 1

    db.session.commit()
    return jsonify({"added": added, "total": len(orders_data)})


@bp.route("/api/orders/<int:order_id>/start", methods=["POST"])
def start_order(order_id):
    order = Order.query.get_or_404(order_id)
    data = request.get_json() or {}
    box_id = data.get("box_id")

    if box_id:
        order.box_id = box_id
    order.status = "picking"
    order.started_at = datetime.utcnow()
    db.session.commit()

    app = current_app._get_current_object()
    thread = threading.Thread(target=_pick_order, args=(app, order.id))
    thread.daemon = True
    thread.start()

    return jsonify({"ok": True, "order": order.to_dict()})


def _pick_order(app, order_db_id: int):
    """Виконати підбір замовлення — активувати серво для кожного товару."""
    with app.app_context():
        order = Order.query.get(order_db_id)
        if not order:
            return

        for item in order.items:
            # Refresh from DB — a concurrent release_servo_item may have already handled this item
            db.session.refresh(item)
            if item.status != "pending":
                continue

            product = Product.query.filter_by(articl=item.articl, is_active=True).first()

            if not product:
                item.status = "manual"
                item.is_manual = True
                db.session.commit()
                continue

            # Каталоговий товар без реальної комірки — ручний збір
            has_cell = product.servo_board == 0x40 and 0 <= product.servo_channel <= 15
            if not has_cell:
                item.status = "manual"
                item.is_manual = True
                db.session.commit()
                continue

            need = item.quantity_ordered - item.quantity_picked
            if need <= 0:
                item.status = "done"
                db.session.commit()
                continue

            available = product.cell_stock

            if available == 0:
                item.status = "manual"
                item.is_manual = True
                db.session.commit()
                continue

            pick_count = min(need, available)
            manual_count = need - pick_count

            servo.release_multiple(product.servo_board, product.servo_channel, pick_count)

            product.cell_stock -= pick_count
            log = StockLog(
                articl=item.articl,
                order_id=order.order_id,
                quantity=pick_count,
            )
            db.session.add(log)
            item.quantity_picked += pick_count
            item.status = "partial" if manual_count > 0 else "done"
            db.session.commit()


@bp.route("/api/orders/<int:order_id>/confirm_manual", methods=["POST"])
def confirm_manual(order_id):
    order = Order.query.get_or_404(order_id)
    data = request.get_json() or {}
    item_id = data.get("item_id")

    if item_id:
        item = OrderItem.query.get(item_id)
        if item:
            item.status = "done"
            # Виставляємо повну кількість — оператор поклав з запасів складу
            # cell_stock НЕ змінюємо, бо товар не з комірки
            item.quantity_picked = item.quantity_ordered
            db.session.commit()

    return jsonify({"ok": True})


@bp.route("/api/orders/<int:order_id>/release_servo_item", methods=["POST"])
def release_servo_item(order_id):
    """Видати конкретну позицію через серво (кнопка 'Конвеєр' для ручних позицій)."""
    order = Order.query.get_or_404(order_id)
    data = request.get_json() or {}
    item_id = data.get("item_id")
    item = next((i for i in order.items if i.id == item_id), None)
    if not item:
        return jsonify({"error": "Позицію не знайдено"}), 404

    product = Product.query.filter_by(articl=item.articl, is_active=True).first()
    if not product:
        return jsonify({"error": "Товар не знайдено в системі"}), 404

    has_cell = product.servo_board == 0x40 and 0 <= product.servo_channel <= 15
    if not has_cell:
        return jsonify({"error": "Товар не прив'язаний до комірки"}), 400

    need = item.quantity_ordered - item.quantity_picked
    if need <= 0:
        return jsonify({"error": "Товар вже зібраний"}), 400

    available = product.cell_stock
    if available <= 0:
        return jsonify({"error": "Комірка порожня"}), 400

    pick_count = min(need, available)
    servo.release_multiple(product.servo_board, product.servo_channel, pick_count)

    product.cell_stock -= pick_count
    item.quantity_picked += pick_count
    item.status = "done" if item.quantity_picked >= item.quantity_ordered else "partial"

    log = StockLog(articl=item.articl, order_id=order.order_id, quantity=pick_count)
    db.session.add(log)
    db.session.commit()
    return jsonify({"ok": True, "picked": pick_count})


@bp.route("/api/orders/<int:order_id>/packed", methods=["POST"])
def pack_order(order_id):
    order = Order.query.get_or_404(order_id)
    order.status = "packed"
    order.packed_at = datetime.utcnow()
    db.session.commit()
    return jsonify({"ok": True})


@bp.route("/api/orders/<int:order_id>/mark_absent", methods=["POST"])
def mark_absent(order_id):
    """Позначити позицію як відсутню (не відібрано, не відвантажується)."""
    order = Order.query.get_or_404(order_id)
    data = request.get_json() or {}
    item_id = data.get("item_id")
    item = next((i for i in order.items if i.id == item_id), None)
    if not item:
        return jsonify({"error": "Позицію не знайдено"}), 404
    item.status = "absent"
    # quantity_picked не скидаємо — серво могло вже видати частину
    db.session.commit()
    return jsonify({"ok": True})


@bp.route("/api/orders/<int:order_id>/done", methods=["POST"])
def done_order(order_id):
    order = Order.query.get_or_404(order_id)
    order.status = "done"
    db.session.commit()
    return jsonify({"ok": True})


@bp.route("/api/orders/<int:order_id>/cancel", methods=["POST"])
def cancel_order(order_id):
    order = Order.query.get_or_404(order_id)
    order.status = "cancelled"
    order.packed_at = datetime.utcnow()
    db.session.commit()
    return jsonify({"ok": True})


@bp.route("/api/orders/<int:order_id>/restore", methods=["POST"])
def restore_order(order_id):
    """Відновити виконане замовлення до стану очікування."""
    order = Order.query.get_or_404(order_id)
    order.status = "pending"
    order.started_at = None
    order.packed_at = None
    for item in order.items:
        item.status = "pending"
        item.quantity_picked = 0
        item.is_manual = False
    db.session.commit()
    return jsonify({"ok": True})


@bp.route("/api/orders/<int:order_id>/add_item", methods=["POST"])
def add_item_to_order(order_id):
    """Додати позицію до замовлення (заміна/доповнення) з обов'язковим коментарем."""
    order = Order.query.get_or_404(order_id)
    data = request.get_json() or {}
    articl = data.get("articl", "").strip()
    try:
        quantity = int(data.get("quantity", 1))
    except (ValueError, TypeError):
        quantity = 1
    comment = data.get("comment", "").strip()
    if not articl:
        return jsonify({"error": "Артикул обов'язковий"}), 400
    if not comment:
        return jsonify({"error": "Коментар обов'язковий"}), 400
    if quantity <= 0:
        return jsonify({"error": "Кількість має бути більше 0"}), 400
    item = OrderItem(
        order_db_id=order.id,
        articl=articl,
        quantity_ordered=quantity,
        is_manual=True,
        status="manual",
        extra_comment=comment,
    )
    db.session.add(item)
    db.session.commit()
    return jsonify({"ok": True, "item": item.to_dict()})


@bp.route("/api/orders/<int:order_id>/comment", methods=["POST"])
def update_order_comment(order_id):
    """Зберегти коментар до замовлення."""
    order = Order.query.get_or_404(order_id)
    data = request.get_json() or {}
    order.comment = data.get("comment", "")
    db.session.commit()
    return jsonify({"ok": True})


@bp.route("/api/orders/<int:order_id>", methods=["DELETE"])
def delete_order(order_id):
    """Видалити замовлення."""
    order = Order.query.get_or_404(order_id)
    db.session.delete(order)
    db.session.commit()
    return jsonify({"ok": True})


# ─── API: Огляд залишків комірок ──────────────────────────────────────────────

@bp.route("/api/cells/overview")
def get_cells_overview():
    """Залишки по комірках vs потреби pending/picking замовлень."""
    cell_products = Product.query.filter(
        Product.servo_board == 0x40,
        Product.servo_channel >= 0,
        Product.servo_channel <= 15,
        Product.is_active == True,
    ).all()

    # Попит по артикулах з усіх необроблених замовлень (pending)
    pending_orders = Order.query.filter(Order.status.in_(["pending"])).all()
    demand = {}
    for order in pending_orders:
        for item in order.items:
            demand[item.articl] = demand.get(item.articl, 0) + item.quantity_ordered

    in_cell_articls = set()
    cells = []
    for p in cell_products:
        needed = demand.get(p.articl, 0)
        cells.append({
            "channel": p.servo_channel,
            "product_id": p.id,
            "articl": p.articl,
            "name": p.name,
            "photo": p.photo,
            "cell_stock": p.cell_stock,
            "cell_capacity": p.cell_capacity,
            "needed": needed,
            "sufficient": p.cell_stock >= needed,
        })
        in_cell_articls.add(p.articl)

    cells.sort(key=lambda x: x["channel"])

    # Товари потрібні в замовленнях але відсутні в комірках
    missing = []
    for articl, needed in demand.items():
        if articl not in in_cell_articls:
            p = Product.query.filter_by(articl=articl).first()
            missing.append({
                "articl": articl,
                "name": p.name if p else "Невідомий товар",
                "needed": needed,
            })
    missing.sort(key=lambda x: x["needed"], reverse=True)

    return jsonify({"cells": cells, "missing": missing})


# ─── API: Товари ──────────────────────────────────────────────────────────────

@bp.route("/api/products")
def get_products():
    products = Product.query.order_by(Product.servo_channel).all()
    return jsonify([p.to_dict() for p in products])


@bp.route("/api/products/by_articl")
def get_product_by_articl():
    """Пошук товару за артикулом (для автозаповнення форми)."""
    articl = request.args.get("q", "").strip()
    if not articl:
        return jsonify({"error": "Порожній артикул"}), 400
    p = Product.query.filter_by(articl=articl).first()
    if p:
        return jsonify(p.to_dict())
    return jsonify({"error": "not found"}), 404


@bp.route("/api/products", methods=["POST"])
def create_product():
    data = request.get_json()
    existing = Product.query.filter_by(
        servo_board=int(data.get("servo_board", 0x40)),
        servo_channel=int(data["servo_channel"])
    ).first()
    if existing:
        return jsonify({"error": f"Канал {data['servo_channel']} вже зайнятий артикулом {existing.articl}"}), 400

    p = Product(
        articl=data["articl"],
        name=data.get("name", ""),
        servo_board=int(data.get("servo_board", 0x40)),
        servo_channel=int(data["servo_channel"]),
        length=float(data.get("length", 0)),
        width=float(data.get("width", 0)),
        height=float(data.get("height", 0)),
        weight=float(data.get("weight", 0)),
        cell_capacity=int(data.get("cell_capacity", 10)),
        cell_stock=int(data.get("cell_stock", 0)),
        sticker_count=int(data.get("sticker_count", 0)),
        sticker_note=data.get("sticker_note", ""),
        comment=data.get("comment", ""),
    )
    db.session.add(p)
    db.session.commit()
    return jsonify(p.to_dict()), 201


@bp.route("/api/products/<int:product_id>", methods=["PUT"])
def update_product(product_id):
    p = Product.query.get_or_404(product_id)
    data = request.get_json()
    for field in ["name", "articl", "length", "width", "height", "weight",
                  "cell_capacity", "cell_stock", "sticker_count",
                  "sticker_note", "comment", "servo_channel", "servo_board", "is_active"]:
        if field in data:
            val = data[field]
            if field in ["length", "width", "height", "weight"]:
                val = float(val)
            elif field in ["cell_capacity", "cell_stock", "sticker_count", "servo_channel", "servo_board"]:
                val = int(val)
            setattr(p, field, val)
    p.updated_at = datetime.utcnow()
    db.session.commit()
    return jsonify(p.to_dict())


@bp.route("/api/products/<int:product_id>", methods=["DELETE"])
def delete_product(product_id):
    """Очистити комірку — видалити товар."""
    p = Product.query.get_or_404(product_id)
    db.session.delete(p)
    db.session.commit()
    return jsonify({"ok": True})


@bp.route("/api/products/<int:product_id>/photo", methods=["POST"])
def upload_photo(product_id):
    p = Product.query.get_or_404(product_id)
    if "photo" not in request.files:
        return jsonify({"error": "Файл не знайдено"}), 400
    file = request.files["photo"]
    if file and allowed_file(file.filename):
        filename = secure_filename(f"product_{product_id}_{file.filename}")
        upload_folder = os.path.join(os.path.dirname(os.path.dirname(__file__)), "static", "uploads")
        os.makedirs(upload_folder, exist_ok=True)
        file.save(os.path.join(upload_folder, filename))
        p.photo = f"/static/uploads/{filename}"
        db.session.commit()
        return jsonify({"photo": p.photo})
    return jsonify({"error": "Невірний формат файлу"}), 400


@bp.route("/api/products/<int:product_id>/photo", methods=["DELETE"])
def delete_photo(product_id):
    """Видалити фото товару."""
    p = Product.query.get_or_404(product_id)
    if p.photo:
        upload_folder = os.path.join(os.path.dirname(os.path.dirname(__file__)), "static", "uploads")
        filename = p.photo.replace("/static/uploads/", "")
        filepath = os.path.join(upload_folder, filename)
        if os.path.exists(filepath):
            os.remove(filepath)
        p.photo = ""
        db.session.commit()
    return jsonify({"ok": True})


@bp.route("/api/products/restock", methods=["POST"])
def restock_product():
    """Поповнити залишок комірки за артикулом або product_id (додати до поточного)."""
    data = request.get_json() or {}
    try:
        quantity = int(data.get("quantity", 0))
    except (ValueError, TypeError):
        quantity = 0
    if quantity <= 0:
        return jsonify({"error": "Кількість має бути більше 0"}), 400

    product_id = data.get("product_id")
    articl = data.get("articl", "").strip()

    if product_id:
        p = Product.query.get(product_id)
    elif articl:
        p = Product.query.filter_by(articl=articl).first()
    else:
        return jsonify({"error": "Вкажіть артикул або product_id"}), 400

    if not p:
        return jsonify({"error": "Товар не знайдено"}), 404
    p.cell_stock = p.cell_stock + quantity
    log = StockLog(articl=p.articl, order_id="", quantity=quantity, log_type="restock")
    db.session.add(log)
    db.session.commit()
    return jsonify({"ok": True, "articl": p.articl, "cell_stock": p.cell_stock})


@bp.route("/api/products/catalog", methods=["POST"])
def create_catalog_product():
    """Створити товар в каталозі без прив'язки до комірки (servo_board=0xFF)."""
    data = request.get_json() or {}
    articl = data.get("articl", "").strip()
    if not articl:
        return jsonify({"error": "Артикул обов'язковий"}), 400
    if Product.query.filter_by(articl=articl).first():
        return jsonify({"error": f"Артикул {articl} вже існує в системі"}), 400
    p = Product(
        articl=articl,
        name=data.get("name", ""),
        servo_board=0xFF,
        servo_channel=-1,
        length=float(data.get("length", 0)),
        width=float(data.get("width", 0)),
        height=float(data.get("height", 0)),
        weight=float(data.get("weight", 0)),
        sticker_count=int(data.get("sticker_count", 0)),
        sticker_note=data.get("sticker_note", ""),
        comment=data.get("comment", ""),
    )
    db.session.add(p)
    db.session.commit()
    return jsonify(p.to_dict()), 201


@bp.route("/api/products/all")
def get_all_products():
    """Всі товари: і ті що в комірках, і ті що тільки в каталозі."""
    products = Product.query.order_by(Product.name).all()
    return jsonify([p.to_dict() for p in products])


@bp.route("/api/products/init_16", methods=["POST"])
def init_16_products():
    articles = [
        "20980", "21411", "22719", "30769", "30770",
        "36401", "36403", "36405", "41276", "41278",
        "41279", "45031", "45035", "45245", "45246", "45742"
    ]
    added = 0
    for i, art in enumerate(articles):
        if Product.query.filter_by(articl=art).first():
            continue
        p = Product(
            articl=art,
            name=f"Товар {art}",
            servo_board=0x40,
            servo_channel=i,
            cell_capacity=20,
            cell_stock=15,
        )
        db.session.add(p)
        added += 1
    db.session.commit()
    return jsonify({"added": added})


# ─── API: Коробки ─────────────────────────────────────────────────────────────

@bp.route("/api/boxes")
def get_boxes():
    boxes = Box.query.filter_by(is_active=True).all()
    return jsonify([b.to_dict() for b in boxes])


@bp.route("/api/boxes", methods=["POST"])
def create_box():
    data = request.get_json()
    b = Box(
        name=data["name"],
        length=float(data["length"]),
        width=float(data["width"]),
        height=float(data["height"]),
        max_fill_pct=float(data.get("max_fill_pct", 80)),
        weight=float(data.get("weight", 0)),
    )
    db.session.add(b)
    db.session.commit()
    return jsonify(b.to_dict()), 201


@bp.route("/api/boxes/<int:box_id>", methods=["PUT"])
def update_box(box_id):
    b = Box.query.get_or_404(box_id)
    data = request.get_json()
    for field in ["name", "length", "width", "height", "max_fill_pct", "weight"]:
        if field in data:
            val = float(data[field]) if field != "name" else data[field]
            setattr(b, field, val)
    db.session.commit()
    return jsonify(b.to_dict())


@bp.route("/api/boxes/<int:box_id>", methods=["DELETE"])
def delete_box(box_id):
    b = Box.query.get_or_404(box_id)
    b.is_active = False
    db.session.commit()
    return jsonify({"ok": True})


@bp.route("/api/boxes/init", methods=["POST"])
def init_boxes():
    defaults = [
        {"name": "XS - 15×10×10", "length": 150, "width": 100, "height": 100},
        {"name": "S - 20×15×10",  "length": 200, "width": 150, "height": 100},
        {"name": "M - 30×20×15",  "length": 300, "width": 200, "height": 150},
        {"name": "L - 40×30×20",  "length": 400, "width": 300, "height": 200},
        {"name": "XL - 50×40×30", "length": 500, "width": 400, "height": 300},
    ]
    added = 0
    for d in defaults:
        if not Box.query.filter_by(name=d["name"]).first():
            db.session.add(Box(**d))
            added += 1
    db.session.commit()
    return jsonify({"added": added})


# ─── API: Звіт ────────────────────────────────────────────────────────────────

def _report_date_range():
    """Повертає (since, until) для фільтрації звіту."""
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")
    if date_from and date_to:
        try:
            since = datetime.strptime(date_from, "%Y-%m-%d")
            until = datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1)
            return since, until
        except ValueError:
            pass
    days = int(request.args.get("days", 7))
    since = datetime.utcnow() - timedelta(days=days)
    return since, None


@bp.route("/api/report")
def get_report():
    since, until = _report_date_range()
    days = int(request.args.get("days", 7))

    q = Order.query.filter(
        Order.started_at >= since,
        Order.status.in_(["packed", "done"])
    )
    if until:
        q = q.filter(Order.started_at < until)
    orders = q.all()

    total_orders = len(orders)
    total_items = sum(len(o.items) for o in orders)
    total_picked = sum(i.quantity_picked for o in orders for i in o.items)
    manual_items = sum(1 for o in orders for i in o.items if i.is_manual)
    not_picked = sum(1 for o in orders for i in o.items if i.status == "absent")

    by_articl = {}
    for o in orders:
        for i in o.items:
            d = i.to_dict()
            if i.articl not in by_articl:
                by_articl[i.articl] = {"articl": i.articl, "name": d["name"], "count": 0}
            by_articl[i.articl]["count"] += i.quantity_picked

    top_products = sorted(by_articl.values(), key=lambda x: x["count"], reverse=True)[:10]

    return jsonify({
        "period_days": days,
        "total_orders": total_orders,
        "total_items": total_items,
        "total_picked": total_picked,
        "manual_items": manual_items,
        "not_picked": not_picked,
        "top_products": top_products,
    })


@bp.route("/api/report/orders")
def get_report_orders():
    """Список замовлень для drill-down у KPI картках."""
    since, until = _report_date_range()
    q = Order.query.filter(
        Order.started_at >= since,
        Order.status.in_(["packed", "done"])
    ).order_by(Order.started_at.desc())
    if until:
        q = q.filter(Order.started_at < until)
    orders = q.all()
    result = []
    for o in orders:
        total_picked = sum(i.quantity_picked for i in o.items)
        manual_count = sum(1 for i in o.items if i.is_manual)
        result.append({
            "id": o.id,
            "order_id": o.order_id,
            "order_date": o.order_date or "",
            "status": o.status,
            "box": o.box.name if o.box else "",
            "started_at": o.started_at.strftime("%d.%m.%Y %H:%M") if o.started_at else "",
            "items_count": len(o.items),
            "total_picked": total_picked,
            "manual_count": manual_count,
        })
    return jsonify(result)


@bp.route("/api/report/items")
def get_report_items():
    """Товари для KPI drill-down: зведення по артикулах за період."""
    since, until = _report_date_range()
    q = Order.query.filter(
        Order.started_at >= since,
        Order.status.in_(["packed", "done"])
    )
    if until:
        q = q.filter(Order.started_at < until)
    orders = q.all()

    by_articl = {}
    for o in orders:
        for i in o.items:
            if i.articl not in by_articl:
                d = i.to_dict()
                by_articl[i.articl] = {
                    "articl": i.articl,
                    "name": d["name"],
                    "picked": 0,
                    "manual": 0,
                }
            by_articl[i.articl]["picked"] += i.quantity_picked
            if i.is_manual:
                by_articl[i.articl]["manual"] += 1

    result = sorted(by_articl.values(), key=lambda x: x["picked"], reverse=True)
    return jsonify(result)


@bp.route("/api/report/absent")
def get_report_absent():
    """Товари з статусом 'absent' (не відібрано) за період."""
    since, until = _report_date_range()
    q = Order.query.filter(
        Order.started_at >= since,
        Order.status.in_(["packed", "done"])
    )
    if until:
        q = q.filter(Order.started_at < until)
    orders_list = q.all()
    rows = []
    for o in orders_list:
        for i in o.items:
            if i.status == "absent":
                d = i.to_dict()
                rows.append({
                    "order_id": o.order_id,
                    "articl": i.articl,
                    "name": d["name"],
                    "quantity_ordered": i.quantity_ordered,
                })
    return jsonify(rows)


@bp.route("/api/report/restock")
def get_report_restock():
    """Поповнення комірок оператором за період."""
    since, until = _report_date_range()
    q = StockLog.query.filter(
        StockLog.log_type == "restock",
        StockLog.created_at >= since,
    )
    if until:
        q = q.filter(StockLog.created_at < until)
    logs = q.order_by(StockLog.created_at.desc()).all()
    by_articl = {}
    for lg in logs:
        p = Product.query.filter_by(articl=lg.articl).first()
        if lg.articl not in by_articl:
            by_articl[lg.articl] = {
                "articl": lg.articl,
                "name": p.name if p else lg.articl,
                "total": 0,
            }
        by_articl[lg.articl]["total"] += lg.quantity
    result = sorted(by_articl.values(), key=lambda x: x["total"], reverse=True)
    # також загальна кількість
    total_restock = sum(lg.quantity for lg in logs)
    return jsonify({"total": total_restock, "items": result})


@bp.route("/api/report/export")
def export_report():
    """Вивантажити звіт у форматі Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({"error": "openpyxl не встановлений"}), 500

    days = int(request.args.get("days", 7))
    since, until = _report_date_range()

    q = Order.query.filter(
        Order.started_at >= since,
        Order.status.in_(["packed", "done"])
    ).order_by(Order.started_at)
    if until:
        q = q.filter(Order.started_at < until)
    orders = q.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Звіт"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    headers = ["Замовлення", "Статус", "Дата комплектації", "Коробка",
               "Артикул", "Назва товару", "Замовлено", "Зібрано серво", "Вручну"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for order in orders:
        date_str = order.started_at.strftime("%d.%m.%Y %H:%M") if order.started_at else ""
        box_name = order.box.name if order.box else ""
        for item in order.items:
            d = item.to_dict()
            ws.append([
                order.order_id,
                order.status,
                date_str,
                box_name,
                item.articl,
                d.get("name", ""),
                item.quantity_ordered,
                item.quantity_picked,
                1 if item.is_manual else 0,
            ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")
    if date_from and date_to:
        filename = f"pick_to_belt_report_{date_from}_{date_to}.xlsx"
    else:
        filename = f"pick_to_belt_report_{days}d.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ─── API: Ручна збірка ────────────────────────────────────────────────────────

@bp.route("/api/manual/channels")
def manual_channels():
    """Повертає всі продукти, прив'язані до каналів серво (для кнопок збірки)."""
    products = Product.query.filter(
        Product.is_active == True,
        Product.servo_board == 0x40,
        Product.servo_channel >= 0,
        Product.servo_channel <= 15
    ).order_by(Product.servo_channel).all()
    return jsonify([{
        "id": p.id,
        "articl": p.articl,
        "name": p.name,
        "channel": p.servo_channel,
        "cell_stock": p.cell_stock,
        "cell_capacity": p.cell_capacity,
        "photo": p.photo,
    } for p in products])


@bp.route("/api/manual/sessions", methods=["GET"])
def manual_sessions_list():
    """Список сесій ручної збірки (остання 50)."""
    sessions = ManualSession.query.order_by(ManualSession.created_at.desc()).limit(50).all()
    return jsonify([s.to_dict() for s in sessions])


@bp.route("/api/manual/sessions", methods=["POST"])
def manual_sessions_create():
    """Створити нову сесію ручної збірки."""
    data = request.get_json() or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Назва події обов'язкова"}), 400
    session = ManualSession(name=name)
    db.session.add(session)
    db.session.commit()
    return jsonify(session.to_dict()), 201


@bp.route("/api/manual/sessions/<int:sid>/close", methods=["POST"])
def manual_sessions_close(sid):
    """Закрити сесію ручної збірки."""
    session = ManualSession.query.get_or_404(sid)
    session.is_closed = True
    session.closed_at = datetime.utcnow()
    db.session.commit()
    return jsonify(session.to_dict())


@bp.route("/api/manual/sessions/<int:sid>/logs")
def manual_session_logs(sid):
    """Логи спрацювань для конкретної сесії."""
    session = ManualSession.query.get_or_404(sid)
    logs = ManualTriggerLog.query.filter_by(session_id=sid).order_by(
        ManualTriggerLog.triggered_at.desc()
    ).all()
    return jsonify({
        "session": session.to_dict(),
        "logs": [l.to_dict() for l in logs]
    })


@bp.route("/api/manual/trigger", methods=["POST"])
def manual_trigger():
    """Спрацювати серво вручну і записати в лог."""
    data = request.get_json() or {}
    session_id = data.get("session_id")
    articl = (data.get("articl") or "").strip()
    quantity = int(data.get("quantity") or 1)

    if not session_id:
        return jsonify({"error": "Потрібно вибрати подію (сесію)"}), 400
    if not articl:
        return jsonify({"error": "Артикул обов'язковий"}), 400
    if quantity < 1:
        quantity = 1

    session_obj = ManualSession.query.get(session_id)
    if not session_obj:
        return jsonify({"error": "Сесію не знайдено"}), 404
    if session_obj.is_closed:
        return jsonify({"error": "Сесія вже закрита"}), 400

    product = Product.query.filter_by(articl=articl, is_active=True).first()
    if not product:
        return jsonify({"error": f"Товар з артикулом {articl} не знайдено"}), 404
    if not (product.servo_board == 0x40 and 0 <= product.servo_channel <= 15):
        return jsonify({"error": "Товар не має прив'язки до серво-каналу"}), 400

    # Спрацювати серво
    try:
        servo.release_multiple(product.servo_board, product.servo_channel, quantity)
    except Exception as e:
        return jsonify({"error": f"Помилка серво: {str(e)}"}), 500

    # Знімаємо залишок
    product.cell_stock = max(0, product.cell_stock - quantity)

    # Записуємо в StockLog
    db.session.add(StockLog(
        articl=articl,
        order_id=f"manual:{session_obj.name}",
        quantity=quantity,
        log_type="pick"
    ))

    # Записуємо в ManualTriggerLog
    log_entry = ManualTriggerLog(
        session_id=session_id,
        articl=articl,
        product_name=product.name,
        channel=product.servo_channel,
        quantity=quantity,
    )
    db.session.add(log_entry)
    db.session.commit()

    return jsonify({
        "ok": True,
        "log": log_entry.to_dict(),
        "cell_stock": product.cell_stock,
        "simulation": servo.is_simulation(),
    })


# ─── API: Тест серво (без запису в лог) ───────────────────────────────────────

@bp.route("/api/admin/test_servo", methods=["POST"])
def admin_test_servo():
    """Калібрування серво — встановити кут і тримати, або повний цикл.
    БЕЗ запису в ManualTriggerLog і StockLog.
    mode='angle': встановити кут angle° і утримувати (для калібрування)
    mode='cycle': повний цикл відкрити→закрити (стандартний тест)
    mode='disable': вимкнути PWM (відпустити серво)
    """
    data = request.get_json() or {}
    channel = data.get("channel")
    mode = data.get("mode", "cycle")

    if channel is None or not (0 <= int(channel) <= 15):
        return jsonify({"error": "Невірний номер каналу"}), 400

    ch = int(channel)
    try:
        if mode == "angle":
            angle = float(data.get("angle", 90))
            angle = max(0.0, min(180.0, angle))
            servo.set_angle(0x40, ch, angle)
        elif mode == "disable":
            servo.disable_channel(0x40, ch)
        else:  # cycle
            servo.release_one(0x40, ch)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    return jsonify({"ok": True, "simulation": servo.is_simulation()})


# ─── API: Налаштування / PIN адміністратора ────────────────────────────────────

def _get_admin_pin():
    """Повертає поточний PIN. Якщо не встановлено — порожній рядок (PIN вимкнений)."""
    s = Settings.query.filter_by(key="admin_pin").first()
    return s.value if s else ""


@bp.route("/api/admin/verify_pin", methods=["POST"])
def admin_verify_pin():
    data = request.get_json() or {}
    pin = (data.get("pin") or "").strip()
    current = _get_admin_pin()
    if not current:
        # PIN не встановлено — доступ вільний
        return jsonify({"ok": True, "pin_required": False})
    return jsonify({"ok": pin == current, "pin_required": True})


@bp.route("/api/admin/pin_required")
def admin_pin_required():
    """Перевіряє чи потрібен PIN взагалі."""
    current = _get_admin_pin()
    return jsonify({"pin_required": bool(current)})


@bp.route("/api/admin/change_pin", methods=["POST"])
def admin_change_pin():
    data = request.get_json() or {}
    current_pin = (data.get("current_pin") or "").strip()
    new_pin = (data.get("new_pin") or "").strip()

    current = _get_admin_pin()
    if current and current_pin != current:
        return jsonify({"error": "Невірний поточний PIN"}), 403

    if new_pin and len(new_pin) < 4:
        return jsonify({"error": "PIN має бути мінімум 4 символи"}), 400

    s = Settings.query.filter_by(key="admin_pin").first()
    if s:
        s.value = new_pin
    else:
        db.session.add(Settings(key="admin_pin", value=new_pin))
    db.session.commit()
    return jsonify({"ok": True, "pin_enabled": bool(new_pin)})


# ─── Статика ──────────────────────────────────────────────────────────────────

@bp.route("/static/uploads/<path:filename>")
def uploaded_file(filename):
    upload_folder = os.path.join(os.path.dirname(os.path.dirname(__file__)), "static", "uploads")
    return send_from_directory(upload_folder, filename)
